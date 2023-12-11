import pandas as pd
import os
import openpyxl
import warnings
import shutil
import psycopg2
from sqlalchemy import create_engine
#from ETL.postgreSQL_config import *
import datetime
import time
import glob
import sys
warnings.simplefilter(action='ignore', category=UserWarning)

# Read Config from env file
import environ

env = environ.Env()
environ.Env.read_env()

dbhost = env('DB_SERVER')
dbport = env('DB_PORT')
dbname = env('DB_NAME_STG')
dbname_reg = env('DB_NAME_REG')
dbname_load = env('DB_NAME_LOAD')
dbuser = env('DB_USER')
dbpassword = env('DB_PASSWORD')
excel_logs_path_ = env('EXCEL_LOG_PATH')
log_path_submission_to_stage = env('LOG_PATH_SUBMISSION_TO_STAGE')
master_log_path = env('MASTER_LOG_PATH')
pre_validation_path_ = env('PRE_VALIDATION_PATH')
to_be_added_to_stage = env('TO_BE_ADDED_TO_STAGE_PATH')
submission_to_stage_path_ = env('SUBMISSION_TO_STAGE_PATH')


def main():

    #Log Functionality(USER):
    df_log=pd.DataFrame() #DataFrame for Log
    file_name_list=[]
    file_path_list=[]
    valid_invalid_list=[]
    current_directory=[]
    datetime_list=[]
    
    #Log Functionality:
    class Logger(object):
        def __init__(self):
            self.terminal = sys.stdout
            self.log = open(f"{log_path_submission_to_stage}submission_to_stage_{datetime.datetime.now()}.log", "a")
    
        def write(self, message):
            self.terminal.write(message)
            self.log.write(message)  

        def flush(self):
            # this flush method is needed for python 3 compatibility.
            # this handles the flush command by doing nothing.
            # you might want to specify some extra behavior here.
            pass

    sys.stdout = Logger()
    time_of_execution=time.strftime("%Y_%m_%d_%H_%M_%S")
    print('Time of Execution:',time_of_execution)

    excel_logs_path = f"{excel_logs_path_}"
    latest_log_folder=max(glob.glob(os.path.join(excel_logs_path, '*/')), key=os.path.getmtime)
    #os.mkdir(user_logs_path)
    
    #Fetching all the Excel files (.xlsx) in the folder
    try:
        #directory=pre_validation_path_
        #latest_folder=max(glob.glob(os.path.join(directory, '*/')), key=os.path.getmtime)
        #latest_folder_extract=f"{latest_folder}VALID/"
        excel_dir = (to_be_added_to_stage).replace('\\','/') # Directory for Excel Files
        latest_folder_2=max(glob.glob(os.path.join(submission_to_stage_path_, '*/')), key=os.path.getmtime)
        successfully_added_to_stage = f"{latest_folder_2}/SUCCESSFULLY_ADDED_TO_STAGE/"
        os.mkdir(successfully_added_to_stage) #Create a Folder to add a copy of submissions added to Stage.
        not_added_to_stage = f"{latest_folder_2}/NOT_ADDED_TO_STAGE/"
        os.mkdir(not_added_to_stage) #Create a Folder to add a copy of submissions that didn't get added to Stage.

    except:
        excel_dir = ''
        print('EXCEL PATH EMPTY')

    ######################################################################
    # STEP-1 : Fetch Tables from DB and get that data inside a Dataframe #
    ######################################################################

    #Fetching Data from template_load_register DB:
    chunk = 10000
    #establishing the connection
    conn = psycopg2.connect(database=dbname_load, user=dbuser, password=dbpassword, host=dbhost, port=dbport)
    # Get json_hash table into a dataframe:
    try:
        df_load_templated_db=pd.read_sql(f"select hashcode from json_hash", con=conn)
        fetched_list=df_load_templated_db['hashcode'].to_list()
        print("json_hash table fetched from DB.")
    except:
        print("json_hash table is absent from template_load_register DB, Setting fetched list as Empty")
        fetched_list=[] #Setting fetched list empty

    ##########################################################################
    # STEP-2 : Fetch Submissions from Blob and compare with the fetched_list #
    ##########################################################################

    #Get all excel files including the subdirectories
    excel_files=[]
    excel_files_with_path=[]

    for path, subdirs, files in os.walk(excel_dir):
        for file in files:
            #print(file)
            if (file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.XLS')):
                temp=os.path.join(path, file)
                print(temp)
                wb = openpyxl.load_workbook(temp)
                meta_created=wb.properties.created
                meta_modified=wb.properties.modified
                meta_sheet_name=[]
                first_five_columns=[]
                json_string=[]
                for name in wb.sheetnames:
                    if name.startswith('Dataset_'):
                        meta_sheet_name.append(name)
                        df_=pd.read_excel(temp,engine='openpyxl',sheet_name=name)
                        df_=df_.dropna(how='all') # Drop Empty Rows
                        df_.columns=df_.columns.str.strip() # Remove Whitespace from Column Names
                        df_.columns=df_.columns.str.lower() # Changing all column names to lower case
                        colm_list=df_.columns.to_list()
                        ff_colm_list=colm_list[:5]
                        first_five_columns.append("".join([str(item) for item in ff_colm_list]))
                        json_string.append(df_.to_json(orient='records'))
                #print(first_five_columns)       
                #hashcode=temp.split("/")[-1]+str(time.mktime(time.strptime(str(meta_created), "%Y-%m-%d %H:%M:%S")))+str(time.mktime(time.strptime(str(meta_modified), "%Y-%m-%d %H:%M:%S")))+str(''.join(meta_sheet_name))+str(''.join(first_five_columns))
                hashcode=''.join(json_string)
                computation=[temp.split("/")[-1],meta_created.strftime('%d/%m/%Y, %H:%M:%S'),meta_modified.strftime('%d/%m/%Y, %H:%M:%S'),meta_sheet_name,first_five_columns,hashcode]
                #print(hashcode)
                if not fetched_list:
                    if computation not in excel_files:
                        excel_files.append(computation)
                        excel_files_with_path.append(temp)
                    else:
                        file_name_list.append(temp.split("/")[-1])
                        file_path_list.append(temp)
                        valid_invalid_list.append('RESUBMISSION')
                        ff_name=temp.split("/")[-1]
                        shutil.move(temp, not_added_to_stage) #Move to NOT_ADDED_TO_STAGE Folder
                        f_name_split=ff_name.split('.') #split filename from the extension
                        new_repeat_f_name=f"{not_added_to_stage}{f_name_split[0]}_RESUBMISSION_ERROR.{f_name_split[-1]}"
                        os.rename(f'{not_added_to_stage}/{ff_name}', new_repeat_f_name)
                        print(f"{ff_name}, Already present in DB, Moved to NOT_ADDED_TO_STAGE folder")
                        #shutil.move(temp_, repetitive_submission_path)
                        #current_directory.append(f"{resubmission_path}{new_repeat_f_name}")
                        datetime_list.append(datetime.datetime.now())
                
                elif fetched_list:
                    if hashcode not in fetched_list:
                        excel_files.append(computation)
                        excel_files_with_path.append(temp)
                    else:
                        file_name_list.append(temp.split("/")[-1])
                        file_path_list.append(temp)
                        valid_invalid_list.append('RESUBMISSION')
                        ff_name=temp.split("/")[-1]
                        shutil.move(temp, not_added_to_stage) #Move to NOT_ADDED_TO_STAGE Folder
                        f_name_split=ff_name.split('.') #split filename from the extension
                        new_repeat_f_name=f"{not_added_to_stage}{f_name_split[0]}_RESUBMISSION_ERROR.{f_name_split[-1]}"
                        os.rename(f'{not_added_to_stage}/{ff_name}', new_repeat_f_name)
                        print(f"{ff_name}, Already present in DB, Moved to NOT_ADDED_TO_STAGE folder")
                        #shutil.move(temp_, repetitive_submission_path)
                        #current_directory.append(f"{resubmission_path}{new_repeat_f_name}")
                        datetime_list.append(datetime.datetime.now())
                else:
                    print("Exception Occured!")

    ##############################################
    # STEP-3 : Updating the Metadata to Database #
    ##############################################

    engine_ = create_engine(f'postgresql://{dbuser}:{dbpassword}@{dbhost}/{dbname_load}'); #template_load_register

    # Adding the Unique Submission Template Metadata to DB (template_load_register):
    try:
        data_f = pd.DataFrame(excel_files)
        data_f.columns = ['submission_name', 'submission_created_at', 'submission_modified_at','sheet_names','first_five_columns','hashcode']
        #data_f.to_csv('submission_template.csv',index=False)
        data_f.to_sql('json_hash', con = engine_, index=False, if_exists='append')
        print('New Submission Template Metadata added to DB')
    except:
        print('No New Data Available.')

    ############################################################
    # STEP-4 : Extracting Required Data from VALID Submissions #
    ############################################################

    #Getting the List of worksheets for each Excel File:
    list_of_worksheets=[] # creating an empty list to store list of lists of workbook names
    for excel in excel_files_with_path:
        try:
            wb = openpyxl.load_workbook(excel)
            list_of_worksheets.append(wb.sheetnames)
        except:
            list_of_worksheets.append('Invalid')
            print('INVALID EXCEL FILE')

    #Getting the List of Worksheet starting with "Dataset_" Prefix:
    filtered_worksheet=[] # creating an empty list to store list of lists of filtered worksheet names
    for i in list_of_worksheets:
        dataset_list = [item for item in i if item.startswith("Dataset_")]
        filtered_worksheet.append(dataset_list)
    print('Filtered Worksheet List:',filtered_worksheet)
    print("EXTRACTION COMPLETED")

    ##############################################################
    # STEP-5 : Transformation and Loading Data to Stage Database #
    ##############################################################

    #Creating Connection Engine for Postgre SQL DB:
    engine = create_engine(f'postgresql://{dbuser}:{dbpassword}@{dbhost}/{dbname}')

    for index,i in enumerate(excel_files_with_path):
        validation_check_list=[]
        for sheet_name in filtered_worksheet[index]:
            df=pd.read_excel(i,sheet_name=sheet_name)
            df=df.dropna(how='all') # Drop Empty Rows
            df.columns=df.columns.str.strip() # Remove Whitespace from Column Names
            df.columns=df.columns.str.lower() # Changing all column names to lower case
            df['time_added_to_db']=datetime.datetime.now() #Adding datetime column
            try:
                #Adding DataFrame to SQL DB:
                df.to_sql(sheet_name, con = engine, chunksize=chunk, method='multi', index=False, if_exists='append')
                print('Table Added:',sheet_name) # Printing the Table Name, after it has been added to the DB
                validation_check_list.append(sheet_name)
            except:
                try:
                    df.rename(columns=lambda x: x[:63], inplace=True)
                    df.to_sql(sheet_name, con = engine, chunksize=chunk, method='multi', index=False, if_exists='append')
                    print('Table Added by limiting column name:',sheet_name) # Printing the Table Name, after it has been added to the DB
                    validation_check_list.append(sheet_name)
                except:
                    try:
                        df_fetched=pd.read_sql(f"select * from \"{sheet_name}\"", con=engine)
                        df_fetched.columns=df_fetched.columns.str.strip() # Remove Whitespace from Column Names
                        df_fetched.columns=df_fetched.columns.str.lower() # Changing all column names to lower case
                        df_fetched.columns=df_fetched.columns.str.replace('"','')
                        df=pd.concat([df_fetched,df], axis=0, ignore_index=True)
                        df.to_sql(sheet_name, con = engine, chunksize=chunk, method='multi', index=False, if_exists='replace')
                        print('Table Fetched and Updated:',sheet_name) # Printing the Table Name, after it has been added to the DB
                        validation_check_list.append(sheet_name)
                    except Exception as m:
                        print('Exception Occured(A):')
                        print(m)

        if set(filtered_worksheet[index]) == set(validation_check_list):
            f_name_v = i.split("/")[-1]
            file_name_list.append(f_name_v)
            file_path_list.append(i)
            valid_invalid_list.append('VALID')
            shutil.move(i, successfully_added_to_stage) #Copy to SUCCESSFULLY_ADDED_TO_STAGE Folder
            print(f"File : {i}, Added to DB and Copied to SUBMISSION_TO_STAGE Folder.")
            #current_directory.append(f"{valid_path}{f_name_v}")
            datetime_list.append(datetime.datetime.now())        
        else:
            missing=(set(filtered_worksheet[index])) - (set(validation_check_list))
            shutil.move(i, not_added_to_stage) #Copy to NOT_ADDED_TO_STAGE Folder
            #print('MISS:',missing)
            #shutil.copy(i, invalid_path) #Copy Invalid File to Folder from Main Directory
            f_name=i.split("/")[-1]
            f_name=f_name.strip()
            file_name_list.append(f_name)
            file_path_list.append(i)
            valid_invalid_list.append('DUPLICATE_COLUMN')
            #current_directory.append(f"{invalid_path}{f_name}")
            datetime_list.append(datetime.datetime.now())
            f_name_split=f_name.split('.') #split filename from the extension
            new_invalid_f_name_with_path=f"{not_added_to_stage}/{f_name_split[0]}_DUPLICATE_COLUMN_ERROR.{f_name_split[-1]}"
            os.rename(f'{not_added_to_stage}/{f_name}', new_invalid_f_name_with_path) #Renaming File name.
            print(f"File : {i}, Didn't get Added to DB.")

    print('ALL Tables Added to DB')

    #Combining Data into Log File
    df_log['File Name']=file_name_list
    df_log['Submission_To_Stage Original Folder Path']=file_path_list
    df_log['Submission_To_Stage Status']=valid_invalid_list
    ec_map =  { 'INVALID' : 'INV',  'VALID' : 'VLD', 'RESUBMISSION' : 'RESUB', 'DUPLICATE_COLUMN': 'DUPCO'}
    df_log['Submission_To_Stage Error Code']=df_log['Submission_To_Stage Status'].map(ec_map)
    #df_log['Current Folder Path']=current_directory
    df_log['Submission_To_Stage Datetime']=datetime_list
    df_log.to_excel(f"{latest_log_folder}/submission_to_stage_user_log.xlsx",index=False)

    # Merging Log File into One:
    df_log_pre_validation=pd.read_excel(f'{latest_log_folder}/prevalidation_user_log.xlsx',engine='openpyxl')
    df_log_submission_to_stage=df_log.copy()

    if df_log_submission_to_stage.empty:
        #df_merge=pd.merge(df_log_pre_validation, df_log_submission_to_stage, on="File Name")
        df_log_pre_validation.to_excel(f"{latest_log_folder}/FINAL_user_log.xlsx",index=False)
        try:
            df_log_pre_validation.to_sql('master_log', con = engine_, index=False, if_exists='append')
        except:
            try:
                df_master_fetched=pd.read_sql(f"select * from master_log", con=engine_)
                df_master_fetched.columns=df_master_fetched.columns.str.replace('"','')
                df_log_pre_validation=pd.concat([df_master_fetched,df_log_pre_validation], axis=0, ignore_index=True)
                df_log_pre_validation.to_sql('master_log', con = engine_, index=False, if_exists='replace')
                print('Master Log Table Fetched and Updated:') # Printing the Table Name, after it has been added to the DB
            except Exception as m_:
                print('Exception Occured(MASTER_LOG):')
                print(m_)
    else:
        df_merge=pd.merge(df_log_pre_validation, df_log_submission_to_stage, on="File Name" , how='outer', suffixes=('_pre_validation', '_submission_to_stage'))
        df_merge=df_merge.fillna('N.A')
        df_merge.to_excel(f"{latest_log_folder}/FINAL_user_log.xlsx",index=False)
        try:
            df_merge.to_sql('master_log', con = engine_, index=False, if_exists='append')
        except:
            try:
                df_master_fetched=pd.read_sql(f"select * from master_log", con=engine_)
                df_master_fetched.columns=df_master_fetched.columns.str.replace('"','')
                df_merge=pd.concat([df_master_fetched,df_merge], axis=0, ignore_index=True)
                df_merge.to_sql('master_log', con = engine_, index=False, if_exists='replace')
                print('Master Log Table Fetched and Updated:') # Printing the Table Name, after it has been added to the DB
            except Exception as m_:
                print('Exception Occured(MASTER_LOG):')
                print(m_)

    df_master_log=pd.read_sql(f"select * from master_log", con=conn)
    df_master_log.to_excel(f"{master_log_path}MASTER_LOG.xlsx",index=False)

if __name__ == '__main__':
    main()