import pandas as pd
import os
import openpyxl
import warnings
import shutil
import psycopg2
from sqlalchemy import create_engine
from numpy import nan
import datetime
import time
import errno
import sys
import re
import glob
from pathlib import PurePath
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
from office365.runtime.auth.client_credential import ClientCredential
from office365.sharepoint.files.file import File
pd.options.mode.chained_assignment = None  # default='warn'
warnings.filterwarnings("ignore")
#warnings.simplefilter(action='ignore', category=UserWarning)

# Read Config from env file
import environ

env = environ.Env()
environ.Env.read_env()

# Database Details:
dbhost = env('DB_SERVER')
dbport = env('DB_PORT')
dbname = env('DB_NAME_STG')
dbname_reg = env('DB_NAME_REG')
dbname_load = env('DB_NAME_LOAD')
dbuser = env('DB_USER')
dbpassword = env('DB_PASSWORD')

# Path Details:
data_directory = env('DATA_DIRECTORY')
remove_empty_folders_from_directory = env('REMOVE_EMPTY_FOLDERS_FROM_DIRECTORY')
log_path_pre_validation = env('LOG_PATH_PRE_VALIDATION')
pre_validation_path_ = env('PRE_VALIDATION_PATH')
to_be_added_to_stage = env('TO_BE_ADDED_TO_STAGE_PATH')
submission_to_stage_path_ = env('SUBMISSION_TO_STAGE_PATH')
excel_logs_path_ = env('EXCEL_LOG_PATH')
prevalidation_master_log_path = env('MASTER_PREVALIDATION_LOG_PATH')
optional_validation= env('OPTIONAL_VALIDATION') #Optional Validation

# Sharepoint Details:
USERNAME = env('sharepoint_email')
PASSWORD = env('sharepoint_password')
SHAREPOINT_SITE = env('sharepoint_url_site')
SHAREPOINT_SITE_NAME = env('sharepoint_site_name')
SHAREPOINT_DOC = env('sharepoint_doc_library')
CLIENT_ID = env('client_id')
CLIENT_SECRET = env('client_secret')

def main():

    #Log Functionality(USER):
    df_log=pd.DataFrame() #DataFrame for Log
    file_name_list=[]
    file_path_list=[]
    valid_invalid_list=[]
    current_directory=[]
    datetime_list=[]

    #Log Functionality(DEV):
    class Logger(object):
        def __init__(self):
            self.terminal = sys.stdout
            self.log = open(f"{log_path_pre_validation}pre_validation_{datetime.datetime.now()}.log", "a")
    
        def write(self, message):
            self.terminal.write(message)
            self.log.write(message)  

        def flush(self):
            # this flush method is needed for python 3 compatibility.
            # this handles the flush command by doing nothing.
            # you might want to specify some extra behavior here.
            pass

    sys.stdout = Logger()
    time_of_execution=time.strftime("%Y_%m_%d_%H_%M_%S")
    print('Time of Execution:',time_of_execution)

    #Fetching all the Excel files (.xlsx) in the folder
    try:
        #/root/airflow/dags/Data/
        #00_MASTER_TEST_SAMPLE/2023/Q1
        excel_dir = (data_directory).replace('\\','/') + '/' # Directory for Excel Files
    except:
        excel_dir = ''
        print('EXCEL PATH EMPTY')

    # Creating INVALID and VALID Folders with Timeframe attached:
    os.mkdir(f"{pre_validation_path_}{time_of_execution}") #create timestamp directory
    os.mkdir(f"{submission_to_stage_path_}{time_of_execution}") #create timestamp directory

    excel_logs_path = f"{excel_logs_path_}{time_of_execution}"
    os.mkdir(excel_logs_path)

    invalid_path = f"{pre_validation_path_}{time_of_execution}/INVALID/"
    os.mkdir(invalid_path) #Create a Folder to store Invalid Files

    resubmission_path = f"{pre_validation_path_}{time_of_execution}/RESUBMISSION/"
    os.mkdir(resubmission_path) #Create a Folder to store Resubmitted Files

    valid_path = f"{pre_validation_path_}{time_of_execution}/VALID/"
    os.mkdir(valid_path) #Create a Folder to store Valid Files
 
    submission_to_stage_path = f"{submission_to_stage_path_}{time_of_execution}/"  # This will be used by next DAG
    #os.mkdir(submission_to_stage_path) #Create a Folder to store SUBMISSION_TO_STAGE Files
    os.makedirs(submission_to_stage_path, exist_ok=True)

    ########################################################################################################
    # STEP-1 : Fetch list from DB to check whether submission already exist or not data inside a Dataframe #
    ########################################################################################################

    # Fetching Data from template_load_register DB:
    # chunk = 10000
    # establishing the connection
    conn = psycopg2.connect(database=dbname_load, user=dbuser, password=dbpassword, host=dbhost, port=dbport)
    # Get submission_templates table into a dataframe:
    try:
        df_load_templated_db=pd.read_sql(f"select hashcode from json_hash", con=conn)
        fetched_list=df_load_templated_db['hashcode'].to_list()
        print("json_hash table fetched from DB.")
        #print('Hashcode List:',fetched_list)
    except:
        print("json_hash table is absent from template_load_register DB, Setting fetched list as Empty")
        fetched_list=[] #Setting fetched list empty

    ########################################################################
    # STEP-2 : Comparing Hashcode of each file from the ones present in DB #
    ########################################################################

    #Get all excel files including the subdirectories
    excel_files=[]
    excel_files_with_path=[]

    for path, subdirs, files in os.walk(excel_dir):
        for file in files:
            #print(file)
            if (file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.XLS')):
                temp=os.path.join(path, file)
                try:
                    wb = openpyxl.load_workbook(temp)
                    meta_created=wb.properties.created
                    meta_modified=wb.properties.modified
                    meta_sheet_name=[]
                    first_five_columns=[]
                    json_string=[]
                    for name in wb.sheetnames:
                        if name.startswith('Dataset_'):
                            meta_sheet_name.append(name)
                            df_=pd.read_excel(temp,engine='openpyxl',sheet_name=name)
                            df_=df_.dropna(how='all') # Drop Empty Rows
                            df_.columns=df_.columns.str.strip() # Remove Whitespace from Column Names
                            df_.columns=df_.columns.str.lower() # Changing all column names to lower case
                            colm_list=df_.columns.to_list()
                            ff_colm_list=colm_list[:5]
                            first_five_columns.append("".join([str(item) for item in ff_colm_list]))
                            json_string.append(df_.to_json(orient='records'))
                    #print(first_five_columns)       
                    #hashcode=temp.split("/")[-1]+str(time.mktime(time.strptime(str(meta_created), "%Y-%m-%d %H:%M:%S")))+str(time.mktime(time.strptime(str(meta_modified), "%Y-%m-%d %H:%M:%S")))+str(''.join(meta_sheet_name))+str(''.join(first_five_columns))
                    hashcode=''.join(json_string)
                    computation=[temp.split("/")[-1],meta_created.strftime('%d/%m/%Y, %H:%M:%S'),meta_modified.strftime('%d/%m/%Y, %H:%M:%S'),meta_sheet_name,first_five_columns,hashcode]
                    file_name_=temp.split("/")[-1]
                    file_name_=file_name_.strip()
                    #print(hashcode)
                    if not fetched_list:
                        if computation not in excel_files:
                            excel_files.append(computation)
                            excel_files_with_path.append(temp)
                            print(f"{file_name_}, Added to List for Processing. [A]")
                        else:
                            file_name_list.append(temp.split("/")[-1])
                            file_path_list.append(temp)
                            valid_invalid_list.append('RESUBMISSION')
                            shutil.move(temp, resubmission_path) #Move Repeating file to Invalid File to Folder
                            f_name_split=file_name_.split('.') #split filename from the extension
                            new_repeat_f_name=f"{resubmission_path}{f_name_split[0]}_RESUBMISSION_ERROR.{f_name_split[-1]}"
                            os.rename(f'{resubmission_path}/{file_name_}', new_repeat_f_name)
                            print(f"{file_name_}, Already present in DB, Moved to RESUBMISSION folder")
                            current_directory.append(f"{new_repeat_f_name}")
                            datetime_list.append(datetime.datetime.now())

                    elif fetched_list:
                        if hashcode not in fetched_list:
                            excel_files.append(computation)
                            excel_files_with_path.append(temp)
                            print(f"{file_name_}, Added to List for Processing. [B]")
                            #print(hashcode)
                        else:
                            file_name_list.append(temp.split("/")[-1])
                            file_path_list.append(temp)
                            valid_invalid_list.append('RESUBMISSION')
                            shutil.move(temp, resubmission_path) #Move Repeating file to Invalid File to Folder
                            f_name_split=file_name_.split('.') #split filename from the extension
                            new_repeat_f_name=f"{resubmission_path}{f_name_split[0]}_RESUBMISSION_ERROR.{f_name_split[-1]}"
                            os.rename(f'{resubmission_path}/{file_name_}', new_repeat_f_name)
                            print(f"{file_name_}, Already present in DB, Moved to RESUBMISSION folder")
                            current_directory.append(f"{new_repeat_f_name}")
                            datetime_list.append(datetime.datetime.now())
                except:
                    shutil.move(temp, invalid_path) #Move Repeating file to Invalid File to Folder
                    print(f"{file} is CORRUPT. Moved to Invalid Folder")
                    corrupt_file=temp.split("/")[-1]
                    corrupt_file_name=corrupt_file.strip()
                    corrupt_file_name_split=corrupt_file_name.split('.') #split filename from the extension
                    new_corrupt_f_name=f"{invalid_path}{corrupt_file_name_split[0]}_CORRUPT.{corrupt_file_name_split[-1]}"
                    os.rename(f'{invalid_path}/{corrupt_file_name}', new_corrupt_f_name)
                    file_name_list.append(temp.split("/")[-1])
                    file_path_list.append(temp)
                    valid_invalid_list.append('CORRUPT')
                    current_directory.append(f"{new_corrupt_f_name}")
                    datetime_list.append(datetime.datetime.now())

            else:
                print(f"Exception Occured for {file}")

    #############################################
    # STEP-3 : Extracting "Dataset_" Worksheets #
    #############################################

    list_of_worksheets=[] # creating an empty list to store list of lists of workbook names
    for excel in excel_files_with_path:
        try:
            wb = openpyxl.load_workbook(excel)
            list_of_worksheets.append(wb.sheetnames)
        except:
            list_of_worksheets.append('Invalid')
            print('INVALID EXCEL FILE')

    #print(list_of_worksheets)

    filtered_worksheet=[] # creating an empty list to store list of lists of filtered worksheet names
    for i in list_of_worksheets:
        dataset_list = [item for item in i if item.startswith("Dataset_")]
        filtered_worksheet.append(dataset_list)

    print('Filtered Worksheet List:',filtered_worksheet)
    print("EXTRACTION COMPLETED")

    #################################
    # STEP-4 : Validation Algorithm #
    #################################

    print('PRE-VALIDATION IN PROGRESS...')
    conn_imp = psycopg2.connect(database=dbname_load, user=dbuser, password=dbpassword, host=dbhost, port=dbport)
    # Get mandatory_fields table into a dataframe:
    imp_fields=pd.read_sql(f"select * from mandatory_fields", con=conn_imp) #fetching from template_load_register DB
    df_error=imp_fields.copy() #copy into new variable
    df_error['error_codes']=list(df_error.groupby(['datafield name']).ngroup()) # creating error_codes list
    error_dict=pd.Series(df_error['error_codes'].values,index=df_error['datafield name']).to_dict() #converting error_codes to dictionary

    # Function to Re-Write Excel Worksheet into the Original File. 
    def write_excel(filename,sheetname,dataframe):
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer: 
            workBook = writer.book
            try:
                workBook.remove(workBook[sheetname])
            except:
                print("Worksheet does not exist")
            finally:
                dataframe.to_excel(writer, sheet_name=sheetname,index=False)
                writer._save()

    # Function to Compute important columns based on optional_validation variable values.
    def compute_important_columns(optional_validation,sheet_name_):
        # IMPORTANT COLUMNS PART:
        if optional_validation == 'yes' or optional_validation == 'Yes' or optional_validation == 'Y' or optional_validation == 'y' :
            #df['time_added_to_db']=datetime.datetime.now() #Adding datetime column
            df_comp=imp_fields.loc[(imp_fields['dataset_code_id'] == sheet_name_) & \
                                (imp_fields['validation for submission deadline'] == 'mandatory field')]
            important_columns_=df_comp['datafield name'].to_list()
            if not important_columns_: #default rules for datasets missing from rules file
                important_columns_=['reportcode','entity_id','entityname','referenceperiodyear','referenceperiod']
        else:
            important_columns_=['reportcode','entity_id','entityname','referenceperiodyear','referenceperiod']
        return(important_columns_)

    # Validation Algorithm:
    for index,i in enumerate(excel_files_with_path):
        validation_check_list=[]
        for sheet_name in filtered_worksheet[index]:
            df=pd.read_excel(i,sheet_name=sheet_name)
            df=df.dropna(how='all') # Drop Empty Rows
            df.columns=df.columns.str.strip() # Remove Whitespace from Column Names
            df.columns=df.columns.str.lower() # Changing all column names to lower case
            #Get important_columns list from the function
            important_columns=compute_important_columns(optional_validation,sheet_name)
            print(f"Important Columns for {sheet_name}:{important_columns}")
            
            # VALIDATION PART:
            try:
                if ((set(important_columns).issubset(df.columns.to_list())==True) and (df[important_columns].isnull().values.any())==False): #Validation Step
                    #Adding DataFrame to SQL DB:
                    #print(important_columns)
                    validation_check_list.append(sheet_name)
                else:
                    print(f"Important Column Missing from: {i} in the following sheet: {sheet_name} ")
                    
            except Exception as e: 
                print('Exception Occured(B):')
                print(e)
                #print("Moved to Invalid:",sheet_name)

        if set(filtered_worksheet[index]) == set(validation_check_list):
            f_name_v = i.split("/")[-1]
            file_name_list.append(f_name_v)
            file_path_list.append(i)
            valid_invalid_list.append('VALID')
            shutil.move(i, valid_path) #Move Valid File to Folder
            print(f"VALID FILE: {i}, Moved.")
            current_directory.append(f"{valid_path}{f_name_v}")
            datetime_list.append(datetime.datetime.now())
        else:
            missing=(set(filtered_worksheet[index])) - (set(validation_check_list))
            shutil.move(i, invalid_path) #Move Invalid File to Folder from Main Directory
            #print('MISS:',missing)
            #shutil.copy(i, invalid_path) #Copy Invalid File to Folder from Main Directory
            f_name=i.split("/")[-1]
            f_name=f_name.strip()
            file_name_list.append(f_name)
            file_path_list.append(i)
            valid_invalid_list.append('INVALID')
            current_directory.append(f"{invalid_path}{f_name}")
            datetime_list.append(datetime.datetime.now())
            f_name_split=f_name.split('.') #split filename from the extension
            new_invalid_f_name_with_path=f"{invalid_path}/{f_name_split[0]}_INVALID_ERROR.{f_name_split[-1]}"
            os.rename(f'{invalid_path}/{f_name}', new_invalid_f_name_with_path) #Renaming File name.

            for sheet_missing in missing:
                sheet_missing=str(sheet_missing).replace('{','').replace('}','').replace("'",'')
                sheet_missing=sheet_missing.strip()
                print('Missing:',sheet_missing)
                file_loc = f"{new_invalid_f_name_with_path}"
                #print([file_loc,sheet_name])
                df_excel=pd.read_excel(file_loc,engine='openpyxl',sheet_name=sheet_missing)
                df_excel=df_excel.dropna(how='all') # Drop Empty Rows
                df_excel.columns=df_excel.columns.str.strip() # Remove Whitespace from Column Names
                df_excel.columns=df_excel.columns.str.lower() # Changing all column names to lower case
                df_col_list=df_excel.columns.to_list()
                important_columns=compute_important_columns(optional_validation,sheet_missing)
                if ((set(important_columns).issubset(df_excel.columns.to_list())==True) and (df_excel[important_columns].isnull().values.any())==True):
                    compute_error_description=[]
                    compute_error_list=[]
                    for idx, r in df_excel[important_columns].iterrows():
                        nulls = list(r[r.isnull()].index)
                        if nulls:
                            error_description=[]
                            converted_to_code=[]
                            for n_value in nulls:
                                error_description.append(n_value)
                                converted_to_code.append(error_dict.get(n_value))
                            compute_error_description.append("-".join([str(item) for item in error_description]))
                            compute_error_list.append("-".join([str(item) for item in converted_to_code]))
                        else:
                            compute_error_description.append('NO_ERROR')
                            compute_error_list.append('NO_ERROR')
                            #df_excel['ERROR_CODE'][idx]=nulls
                    df_excel['ERROR']=compute_error_list
                    df_excel['ERROR_DESC']=compute_error_description
                    df_excel['NULL']=df_excel[important_columns].isnull().sum(axis=1)
                    df_excel['ERROR_CODE']=df_excel['ERROR'].astype(str)+'::'+df_excel['NULL'].astype(str)
                    df_excel['ERROR_DESCRIPTION']='MISSING_COLUMNS:'+df_excel['ERROR_DESC'].astype(str)+'::'+'NULL_VALUES:'+df_excel['NULL'].astype(str)
                    df_excel.drop(['ERROR','ERROR_DESC' , 'NULL'], axis=1, inplace=True)
                    write_excel(file_loc,sheet_missing,df_excel)
                else: #((set(important_columns).issubset(df_excel.columns.to_list())==False)):
                    #important_columns=compute_important_columns(optional_validation,sheet_missing)
                    #print(important_columns)
                    missing_cols=list(set(important_columns).difference(df_col_list))
                    print('Missing_Important_Columns:',missing_cols)
                    for current_missing_col in missing_cols:
                        df_excel[current_missing_col]=nan
                    #print('Missing Important Columns:',missing_cols)
                    compute_error_description=[]
                    compute_error_list=[]
                    for idx, r in df_excel[important_columns].iterrows():
                        nulls = list(r[r.isnull()].index)
                        if nulls:
                            error_description=[]
                            converted_to_code=[]
                            for n_value in nulls:
                                error_description.append(n_value)
                                converted_to_code.append(error_dict.get(n_value))
                            compute_error_description.append("-".join([str(item) for item in error_description]))
                            compute_error_list.append("-".join([str(item) for item in converted_to_code]))
                        else:
                            compute_error_description.append('NO_ERROR')
                            compute_error_list.append('NO_ERROR')
                            #df_excel['ERROR_CODE'][idx]=nulls
                    df_excel['ERROR']=compute_error_list
                    df_excel['ERROR_DESC']=compute_error_description
                    df_excel['NULL']=df_excel[important_columns].isnull().sum(axis=1)
                    df_excel['ERROR_CODE']=df_excel['ERROR'].astype(str)+'::'+df_excel['NULL'].astype(str)
                    df_excel['ERROR_DESCRIPTION']='MISSING_COLUMNS:'+df_excel['ERROR_DESC'].astype(str)+'::'+'NULL_VALUES:'+df_excel['NULL'].astype(str)
                    df_excel.drop(['ERROR','ERROR_DESC' , 'NULL'], axis=1, inplace=True)
                    df_excel.drop(missing_cols, axis=1, inplace=True)
                    write_excel(file_loc,sheet_missing,df_excel)

                if '_ERROR' not in sheet_missing:
                    ss = openpyxl.load_workbook(file_loc)
                    ss_sheet = ss[f"{sheet_missing}"]
                    ss_sheet.title = f"{sheet_missing}_ERROR"
                    ss.save(file_loc)

            print(f"INVALID FILE: {i}, Moved.")

    # Sending a COPY of VALID Folder Data to TO_BE_LOADED_TO_STAGE Folder:
    # path to source directory
    src_dir = valid_path
    
    # path to destination directory
    dest_dir = to_be_added_to_stage
    
    # getting all the files in the source directory
    files = os.listdir(src_dir)
    shutil.copytree(src_dir, dest_dir, dirs_exist_ok=True)

    print('VALID Folder Data Copied to TO_BE_LOADED_TO_STAGE Folder')

    # Combining Data into Log File

    df_log['File Name']=file_name_list
    df_log['Original Folder Path']=file_path_list
    df_log['Pre-Validation Status']=valid_invalid_list
    ec_map =  { 'INVALID' : 'INV',  'VALID' : 'VLD', 'CORRUPT' : 'CRPT' , 'RESUBMISSION' : 'RESUB'}
    df_log['Error Code']=df_log['Pre-Validation Status'].map(ec_map)
    df_log['Current Folder Path']=current_directory
    df_log['Datetime Prevalidation']=datetime_list
    df_log.to_excel(f"{excel_logs_path}/prevalidation_user_log.xlsx",index=False)

    engine_ = create_engine(f'postgresql://{dbuser}:{dbpassword}@{dbhost}/{dbname_load}'); #template_load_register

    try:
        df_log.to_sql('prevalidation_log', con = engine_, index=False, if_exists='append')
    except:
        try:
            df_log_fetched=pd.read_sql(f"select * from prevalidation_log", con=engine_)
            df_log_fetched.columns=df_log_fetched.columns.str.replace('"','')
            df_log=pd.concat([df_log_fetched,df_log], axis=0, ignore_index=True)
            df_log.to_sql('prevalidation_log', con = engine_, index=False, if_exists='replace')
            print('Prevalidation Log Table Fetched and Updated:') # Printing the Table Name, after it has been added to the DB
        except Exception as m_:
            print('Exception Occured(PREVALIDATION_MASTER_LOG):')
            print(m_)

    df_master_log_prevalidation=pd.read_sql(f"select * from prevalidation_log", con=conn)
    df_master_log_prevalidation.to_excel(f"{prevalidation_master_log_path}PREVALIDATION_MASTER_LOG.xlsx",index=False)

    ###########################################################
    # STEP-5 : Copy VALID Files to SUBMISSION_TO_STAGE Folder #
    ###########################################################

    # #MOVING VALID FILES TO SUBMISSION_TO_STAGE Folder:
    # # Source path
    # src = valid_path
    # # Destination path
    # dest = submission_to_stage_path

    # # Copy the content of VALID folder to SUBMISSION_TO_STAGE folder:
    # try:
    #     shutil.copytree(src, dest, dirs_exist_ok = True)
    #     print('Files SUCCESSFULLY Copied From VALID Folder to SUBMISSION_TO_STAGE Folder!')
    # except OSError as err:
    #     # error caused if the source was not a directory
    #     if err.errno == errno.ENOTDIR:
    #         shutil.copy2(src, dest)
    #     else:
    #         print("Error: % s" % err)

    #########################################################
    # STEP-6 : UPLOADING Invalid Files to Sharepoint Folder #
    #########################################################
    
    client_credentials = ClientCredential(client_id=f'{CLIENT_ID}', client_secret=f'{CLIENT_SECRET}')

    # Defining Class for SharePoint:
    class SharePoint:
        def _auth(self):
            conn = ClientContext(SHAREPOINT_SITE).with_credentials(
                    client_credentials
            )
            return conn
        
        def _get_files_list(self, folder_name):
            conn = self._auth()
            target_folder_url = f'{SHAREPOINT_DOC}/{folder_name}'
            root_folder = conn.web.get_folder_by_server_relative_url(target_folder_url)
            root_folder.expand(["Files", "Folders"]).get().execute_query()
            return root_folder.files
        
        def get_folder_list(self, folder_name):
            conn = self._auth()
            target_folder_url = f'{SHAREPOINT_DOC}/{folder_name}'
            root_folder = conn.web.get_folder_by_server_relative_url(target_folder_url)
            root_folder.expand(["Folders"]).get().execute_query()
            return root_folder.folders

        def download_file(self, file_name, folder_name):
            conn = self._auth()
            file_url = f'/sites/{SHAREPOINT_SITE_NAME}/{SHAREPOINT_DOC}/{folder_name}/{file_name}'
            file = File.open_binary(conn, file_url)
            return file.content
        
        def download_latest_file(self, folder_name):
            date_format = "%Y-%m-%dT%H:%M:%SZ"
            files_list = self._get_files_list(folder_name)
            file_dict = {}
            for file in files_list:
                dt_obj = datetime.datetime.strptime(file.time_last_modified, date_format)
                file_dict[file.name] = dt_obj
            # sort dict object to get the latest file
            file_dict_sorted = {key:value for key, value in sorted(file_dict.items(), key=lambda item:item[1], reverse=True)}    
            latest_file_name = next(iter(file_dict_sorted))
            content = self.download_file(latest_file_name, folder_name)
            return latest_file_name, content
            
        def upload_file(self, file_name, folder_name, content):
            conn = self._auth()
            target_folder_url = f'/sites/{SHAREPOINT_SITE_NAME}/{SHAREPOINT_DOC}/{folder_name}'
            target_folder = conn.web.get_folder_by_server_relative_path(target_folder_url)
            response = target_folder.upload_file(file_name, content).execute_query()
            return response
        
        def upload_file_in_chunks(self, file_path, folder_name, chunk_size, chunk_uploaded=None, **kwargs):
            conn = self._auth()
            target_folder_url = f'/sites/{SHAREPOINT_SITE_NAME}/{SHAREPOINT_DOC}/{folder_name}'
            target_folder = conn.web.get_folder_by_server_relative_path(target_folder_url)
            response = target_folder.files.create_upload_session(
                source_path=file_path,
                chunk_size=chunk_size,
                chunk_uploaded=chunk_uploaded,
                **kwargs
            ).execute_query()
            return response
        
        def get_list(self, list_name):
            conn = self._auth()
            target_list = conn.web.lists.get_by_title(list_name)
            items = target_list.items.get().execute_query()
            return items
            
        def get_file_properties_from_folder(self, folder_name):
            files_list = self._get_files_list(folder_name)
            properties_list = []
            for file in files_list:
                file_dict = {
                    'file_id': file.unique_id,
                    'file_name': file.name,
                    'major_version': file.major_version,
                    'minor_version': file.minor_version,
                    'file_size': file.length,
                    'time_created': file.time_created,
                    'time_last_modified': file.time_last_modified
                }
                properties_list.append(file_dict)
                file_dict = {}
            return properties_list

    # Defining Path for INVALID and RESUBMISSION Files & Upload to SharePoint:

    ROOT_DIR_INVALID = invalid_path
    ROOT_DIR_RESUBMISSION = resubmission_path
    SHAREPOINT_FOLDER_NAME = env('INVALID_SHAREPOINT_FOLDER_NAME')
    SHAREPOINT_FOLDER_NAME_RESUBMISSION = env('RESUBMISSION_SHAREPOINT_FOLDER_NAME')
    FILE_NAME_PATTERN = env('FILE_EXTENSION_INVALID')

    def get_list_of_files(folder):
        file_list = []
        folder_item_list = os.listdir(folder)
        for item in folder_item_list:
            item_full_path = PurePath(folder, item)
            if os.path.isfile(item_full_path):
                file_list.append([item, item_full_path])
        return file_list

    # read files and return the content of files
    def get_file_content(file_path):
        with open(file_path, 'rb') as f:
            return f.read()
        
    def upload_files(folder, keyword=None):
        file_list = get_list_of_files(folder)
        for file in file_list:
            if keyword is None or keyword == 'None' or re.search(keyword, file[0]):
                file_content = get_file_content(file[1])
                SharePoint().upload_file(file[0], SHAREPOINT_FOLDER_NAME, file_content)

    def upload_files_resubmission(folder, keyword=None):
        file_list = get_list_of_files(folder)
        for file in file_list:
            if keyword is None or keyword == 'None' or re.search(keyword, file[0]):
                file_content = get_file_content(file[1])
                SharePoint().upload_file(file[0], SHAREPOINT_FOLDER_NAME_RESUBMISSION, file_content)

    # Upload Part:
    upload_files(ROOT_DIR_INVALID, FILE_NAME_PATTERN) # Upload INVALID Files to SharePoint
    print('Invalid Files Uploaded to Sharepoint!')
    upload_files_resubmission(ROOT_DIR_RESUBMISSION, FILE_NAME_PATTERN) # Upload RESUBMISSION Files to SharePoint
    print('Resubmission Files Uploaded to Sharepoint!')
    
    def delete_empty_folders(root):
        for dirpath, dirnames, filenames in os.walk(root, topdown=False):
            for dirname in dirnames:
                full_path = os.path.join(dirpath, dirname)
                if not os.listdir(full_path): 
                    os.rmdir(full_path)

    # Removing empty Folders inside the Directory:
    delete_empty_folders(remove_empty_folders_from_directory)
                         
if __name__ == '__main__':
    main()
